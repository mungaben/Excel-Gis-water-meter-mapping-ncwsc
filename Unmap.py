import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Not.xlsx')

# Select the worksheet
ws=wb.worksheets[0]
print(ws.title)
# ws = wb.active

# Get the data range
data_range = ws.iter_rows(values_only=True)
print(data_range)

# Sort the data by the first column
sorted_data = sorted(data_range, key=lambda x: x[1])

# yx=[x for x[17] in data_range if x is not None]
# print(len(yx))
# for x in data_range:
   

# sorted_data = sorted([x for x in data_range if x[17] is not None], key=lambda x: x[17])

print(len(sorted_data[0]))
title=sorted_data[0]

# # Create a dictionary to store data for each unique value in the first column
data_dict = {}

# # Iterate over the sorted data and group it by the first column value
for row in sorted_data:
    first_col_value = row[0]

    if first_col_value not in data_dict:
        # print('First column value not in dictionary', first_col_value)
        data_dict[first_col_value] = []

    data_dict[first_col_value].append(row)
# print("data dictionary",data_dict)
# # Create a separate file for each unique value in the first column
new_wb = openpyxl.Workbook()

# # col_indices = {n: cell.value for n, cell in enumerate(new_sheet.rows[0]) 
# #                if cell.value in title}
# # print(col_indices)

for value, data in data_dict.items():
    # Create a new workbook and worksheet
    # print('Creating new workbook and worksheet for',value, data[0])
    new_sheet = new_wb.active
   
#     # Write the data to the new worksheet
    for row_index, row_data in enumerate(data, start=1):
        
        # new_sheet.append(title)
        # for row_value0, row_title1 in enumerate(title, start=1):
            # print('title data a to new worksheet',  row_title1)
            # new_sheet.cell(row=row_index, column=row_value0, value=row_title1)
        
        for col_index, value in enumerate(row_data, start=1):
            # print('Writing data to new worksheet',  value)
            
            
            new_sheet.cell(row=row_index, column=col_index, value=value)
            # print('Writing data to new worksheet',  new_sheet)

#     # Save the new workbook with the value from the first column as the filename
# print('Saving new workbook',new_sheet)
    new_sheet.title=data[0][0]
#     # wb.save((f'{data[0][17]}.xlsx'))
    new_wb.save(f'{data[0][0]}.xlsx')
