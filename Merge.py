from openpyxl import Workbook, load_workbook


# Load the first sheet
wb = load_workbook('datas8.xlsx')
print(wb.sheetnames)
sheet1 = wb.worksheets[4]
# print(sheet1)

# # Load the second sheet
sheet2 = wb.worksheets[1]
print(sheet2)


merged_sheet = wb.create_sheet(title='Merged')
print(merged_sheet)

sheet1data=[]
for row in sheet1.iter_rows(values_only=True):
    sheet1data.append(row)
    print("sheet1",row)
    
    # print(row)
    
    
sheet2data=[]
for row in sheet2.iter_rows(values_only=True):
    sheet2data.append(row)
    print("sheet2",row)
    
alldata=[]
for data2 in sheet2data:
    for data in sheet1data:
        if data[1] == data2[2]:
            print("data1",data)
            print("data2",data2)
            # print("data",data[3])
#             # Merge the data from the two sheets
           
#             # Convert the tuple to a list
            data_list = list(data2)

# # # # Add the new value to the list
            data_list.append(data[2])   

# # # # Convert the list back to a tuple
            updated_tuple = tuple(data_list)

# #             # print(updated_tuple)
            alldata.append(updated_tuple)
# #             # print(merged_sheet)
            
            
print("dataall",alldata)
            
 



# for row in sheet2.iter_rows(min_row=2, values_only=True):
#     first_column_value = row[2]
    
#     found = False
#     print(first_column_value)

# #     # Search for a matching value in the first column of the merged sheet
#     for merged_row in merged_sheet.iter_rows(min_row=2, values_only=True):
#         print(merged_row)
        
#         if merged_row[2] == first_column_value:
#             # print(merged_row[3])
#             merged_row[2]
#             found = True
# #             # Update the merged sheet with the additional data from the second sheet
#             merged_row[2] = row[1]  # Assuming the additional data is in the second column

# # #     # If no match is found, append a new row to the merged sheet with the data from the second sheet
#     if not found:
#         merged_sheet.append(row)


wb2 = Workbook()


sheet = wb.active


for row_index, data_row in enumerate(alldata, start=1):
    for col_index, value in enumerate(data_row, start=1):
        sheet.cell(row=row_index, column=col_index, value=value)


wb.save('datas9.xlsx')



