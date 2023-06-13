

# creates a wookbook shets from a single workbook
import openpyxl

# Load the source workbook
source_workbook = openpyxl.load_workbook('UNMAPED.xlsx')

# Iterate over the sheets in the source workbook
for sheet_name in source_workbook.sheetnames:
    # Create a new workbook
    new_workbook = openpyxl.Workbook()
    
    # Get the source sheet
    source_sheet = source_workbook[sheet_name]
    
    # Get the active sheet of the new workbook
    new_sheet = new_workbook.active
    
    # Copy the source sheet data to the new sheet
    for row in source_sheet.iter_rows(values_only=True):
        new_sheet.append(row)
    
    # Save the new workbook with the sheet name as the filename
    new_workbook.save(f'{sheet_name}.xlsx')

# Close the source workbook
source_workbook.close()
