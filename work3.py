from openpyxl import load_workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Load the original workbook
original_workbook = load_workbook('Not.xlsx')
original_sheet = original_workbook.worksheets[0]
print(original_sheet.title)

# Get the data range to sort
data_range = original_sheet.iter_rows(values_only=True)
print(data_range)

# Sort the data by column 17
# sorted_data = sorted(data_range, key=lambda x: x[0])
# print(sorted_data)


sorted_data = sorted([x for x in data_range if x[0] is not None], key=lambda x: x[0])
print(sorted_data[0])


data_dict = {}
titles=[]
for row in sorted_data:
    first_col_value = row[0]
    # titles.append(row[0])
   
    
    # print('First column value row', row)
    # print('First column value', first_col_value)
# push data to data dictionary if not there also add it
    if first_col_value not in data_dict:
        # print('First column value not in dictionary', first_col_value)
        data_dict[first_col_value] = []

    data_dict[first_col_value].append(row)
print("data dictionary",len(data_dict))


# # Create a new workbook
new_workbook = load_workbook()
sheet_index = 1

for row in sorted_data:
    # Create a new sheet for each sorted row
    new_sheet = new_workbook.create_sheet(title=str(row[17]))

    # Add the header from the original sheet
    header_row = list(original_sheet.iter_rows(max_row=1, values_only=True))
    new_sheet.append(header_row[0])

    # Add the sorted row to the new sheet
    new_sheet.append(row)

    # Apply bold font to the header row
    # for cell in new_sheet[1]:
    #     cell.font = Font(bold=True)

    # Increment the sheet index for the next iteration
    sheet_index += 1

# # Remove the default sheet created with the new workbook
# new_workbook.remove(new_workbook.active)

# # Save the new workbook
new_workbook.save('sorted_files.xlsx')

# # Create a new workbook
new_workbook =openpyxl.Workbook()
# sheet_index = 1

# # # Create a separate file for each unique value in the first column
# for value, data in data_dict.items():
#     # # Create a new workbook and worksheet
#     # print('Creating new workbook and worksheet for', data[0])
#     # new_wb = openpyxl.Workbook()
#     # new_sheet = new_wb.active
    
#     # Create a new sheet for each sorted row
#     new_sheet = new_workbook.create_sheet(title=str(data[0][17]))
#     print(new_sheet.title)
    
        
#     # print('sheet data',new_sheet[1])

# #     # Write the data to the new worksheet
#     for row_index, row_data in enumerate(data, start=1):
#         new_sheet.append(sorted_data[0])
        
#         # for cell in new_sheet[1]:
#         #     cell.font = Font(bold=True)
#         #     print('cell',cell.value)
   
#         for col_index, value in enumerate(row_data, start=1):
#             # print('Writing data to new worksheet',  value)
#             new_sheet.cell(row=row_index, column=col_index, value=value)
    
# # #     # Save the new workbook with the value from the first column as the filename
# # for cell in new_sheet[1]:
# #         # cell.font = Font(bold=True)
# #         print('cell',cell.value)

# # new_workbook.remove(new_workbook.active)
# # new_workbook.save("mapped&ux.xlsx")




