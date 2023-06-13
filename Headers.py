
from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook('themappeddata.xlsx')

# Get the list of sheet names
sheet_names = workbook.sheetnames

# Iterate over each sheet
for sheet_name in sheet_names:
    # Get the sheet by name
    sheet = workbook[sheet_name]

    # Create a list of header values
    headers = ['Installation key', 'Installation Type', 'Customer Owner id', 'Customer Name', 'Tenant Owner id', 'Tenant Name', 'Inst Region', 'Inst Activity type', 'Inst No of Dwelling units', 'Inst Landmark', 'Inst Zone', 'Inst Sub Zone', 'Inst Area', 'Inst Street', 'Inst House Number', 'Inst LR Number', 'Inst Category', 'Inst Route', 'Inst Walk Order', 'Inst Tarrif', 'Inst Supply type', 'Meter Number', 'Meter Installation Date', 'Smart Meter', 'Test Meter', 'Meter Last Reading', 'Meter Last Reading Date', 'Meter Last Billed Date', 'Meter Last Billed reading', 'Meter Multiply factor', 'Meter Number of Dials', 'Meter Usage', 'Meter Shadow', 'Meter Make', 'Meter Size', 'Meter Placement', 'Meter Status', 'Meter Connection status', 'ObjectId', 'GlobalID', 'Installation Balance Date', 'Installation Survey By', 'Installation Survey Date', 'CreationDate', 'Creator', 'EditDate', 'Editor', 'Location Update Date', 'Installation Balance', 'Plot Number', 'Apartment Name', 'No of Floors', 'Comments', 'No of Units', 'Surveyor PF', 'srv Inst Tarrif', 'srv Inst Supply Type', 'x', 'y']

    # Insert the headers at the beginning of the sheet
    sheet.insert_rows(1)
    for col_num, header in enumerate(headers, start=1):
        # quotient, remainder = divmod(col_num, 26)
        # if remainder == 0:
        #     col_letter = chr(64 + quotient - 1) + 'Z'
        # else:
        #     col_letter = chr(64 + quotient) + chr(64 + remainder)
        if col_num <= 26:
            col_letter = chr(64 + col_num)
        else:
            quotient, remainder = divmod(col_num, 26)
            if remainder == 0:
                col_letter = chr(64 + quotient - 1) + 'Z'
            else:
                col_letter = chr(64 + quotient) + chr(64 + remainder)
        cell_coordinate = f"{col_letter}1"
        print("cell cordinates", cell_coordinate)
        print(f"Inserting header {header} at {cell_coordinate}")
        sheet[cell_coordinate].value = header

# Save the modified workbook
workbook.save('mappedData.xlsx')
