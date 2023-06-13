import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# Load the Excel file
filename = 'Note.xlsx'
workbook = openpyxl.load_workbook(filename)

# Select the active sheet
sheet = workbook.active

# Get the data range
data_range = sheet['A1': f'{get_column_letter(sheet.max_column)}{sheet.max_row}']


# Sort the data by the first column
sorted_data = sorted(data_range, key=lambda x: x[0].value)

# Create a new workbook for sorted data
sorted_workbook = Workbook()
sorted_sheet = sorted_workbook.active

# Write the sorted data to the new workbook
for row in sorted_data:
    for cell in row:
        sorted_sheet[cell.coordinate].value = cell.value

# Save the sorted workbook with a new filename
sorted_filename = 'sorted_' + filename
sorted_workbook.save(sorted_filename)
