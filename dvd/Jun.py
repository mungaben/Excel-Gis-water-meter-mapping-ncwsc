import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Note.xlsx')

# Select the worksheet
ws = wb.active
print(ws)

# Get the data range
data_range = ws.iter_rows(values_only=True)
print(data_range)

# Sort the data by the first column
sorted_data = sorted(data_range, key=lambda x: x[0])
# print(sorted_data)

# # # Create a dictionary to store data for each unique value in the first column
data_dict = {}

# # # Iterate over the sorted data and group it by the first column value
for row in sorted_data:
    first_col_value = row[0]

    if first_col_value not in data_dict:
        print('First column value not in dictionary', first_col_value)
        data_dict[first_col_value] = []

    data_dict[first_col_value].append(row)
print("data dictionary",len(data_dict))
# # # Create a separate file for each unique value in the first column
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active 
for value, data in data_dict.items():
    # Create a new workbook and worksheet
    print('Creating new workbook and worksheet for', data[0][0])
    print('Creating sheet data length', len(data))
    
    

    # Write the data to the new worksheet
    for row_index, row_data in enumerate(data, start=1):
        print('Writing data to new worksheet', len(data))
        
        for col_index, value in enumerate(row_data, start=1):
            print('Writing data to new worksheet',  value)
            new_sheet.cell(row=row_index, column=col_index, value=value)

    # # Save the new workbook with the value from the first column as the filename
    new_wb.save(f'{data[0][0]}.xlsx')
