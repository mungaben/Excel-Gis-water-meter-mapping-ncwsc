import openpyxl

wb = openpyxl.load_workbook('tasks.xlsx')
ws=wb.worksheets[0]
Itinerary_Number=[]
for row in ws:
    name=row[1].value
    # check if name starts wiyh  NER 1st letter
    if name and name.startswith("NER_"):
        Itinerary_Number.append(name)
sheet=wb.worksheets[2]
# print(sheet)
install_routes=[]
ws_output=[]

for row in sheet:
    inst_route=row[17].value
    # print(inst_route)
    install_routes.append(inst_route)
    for filterddata in Itinerary_Number:
        # check if filtered data is intalled routes
        if filterddata is inst_route:
            # get the whole data for the filterddata
            
            # row that has the value /filterddata
            
            instat_tarrif=row[56].value
            # print("instat_tarrif ",instat_tarrif)
            
            if instat_tarrif is None:
                # print("instat_tarrif is None")
                meternumber=row[21].value
                # print("meternumber ",meternumber)
                accountNumber=row[36].value
                # print("accountNumber ",accountNumber)
                
                
                # wb.create_sheet(title="Output")
                # ws_output.append(["filterddata,","Meter Number", "Account Number"])
                ws_output.append([filterddata,meternumber, accountNumber])
                
                



print(ws_output)
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Write the data to the new worksheet
for row_index, row_data in enumerate(ws_output, start=1):
    for col_index, value in enumerate(row_data, start=1):
        new_sheet.cell(row=row_index, column=col_index, value=value)

# Rename the new worksheet
new_sheet.title = "Output Data"

# Save the new workbook
new_wb.save('filtered_datas.xlsx')



