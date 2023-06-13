import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Nott.xlsx')

# Select the worksheet
ws = wb.worksheets[1]

# Get the data range
data_range = ws.iter_rows(values_only=True)

# Sort the data by the 17th column
sorted_data = sorted(data_range, key=lambda x: x[0])

# Create a dictionary to store data for each unique value in the 17th column
data_dict = {}

# Iterate over the sorted data and group it by the 17th column value
for row in sorted_data:
    col_value = row[0]  # 17th column since Python uses 0-based indexing

    if col_value not in data_dict:
        data_dict[col_value] = []

    data_dict[col_value].append(row)
new_wb = openpyxl.Workbook()
# Create a separate file for each unique value in the 17th column
for value, data in data_dict.items():
    # Create a new workbook and worksheet
    print(value)
   
    new_sheet = new_wb.active

    # Set the title as the row number 17
    title_row = 0
    title_data = data[0]  # Assuming the title data is the same for all rows with the same value in the 17th column
    

    # Write the title data to the new worksheet
    for col_index, value in enumerate(title_data, start=1):
        new_sheet.cell(row=title_row, column=col_index, value=value)
        # print(new_sheet)

    # Write the remaining data to the new worksheet
    for row_index, row_data in enumerate(data, start=title_row + 1):
        for col_index, value in enumerate(row_data, start=1):
            new_sheet.cell(row=row_index, column=col_index, value=value)

    # # Save the new workbook with the value from the 17th column as the filename
    # new_wb.save(f'{value}.xlsx')
    new_sheet.title=data[0][0]

# Save the new workbook
new_wb.save('Notx.xlsx')