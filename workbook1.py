import openpyxl

wb = openpyxl.load_workbook('Book3.xlsx')
# first sheet
ws=wb.worksheets[0]
print(ws)
Itinerary_Number=[]
for row in ws:
    name=row[17].value
    # print(name)
    # check if name starts wiyh  NER 1st letter
    if name and name.startswith("NER_"):
        Itinerary_Number.append(name)
        
        
        

sheet=wb.worksheets[0]
print(sheet)
ColName = []

# Iterate over columns and retrieve names
for column in sheet.iter_cols(1, sheet.max_column):
    column_name = column[0].value
    print(column_name)
    ColName.append(column_name)
    
new_wb = openpyxl.Workbook()
Install_Tarrifs=[]
ws_output=[]

for row in sheet:
    inst_route=row[17].value
    Install_Tarrif=row[56].value
    # print("inst_route  row data column ",row[0].value)
    
    # print("number_install Trrif",Install_Tarrif)
    # install_routes.append()
    if Install_Tarrif is None:
        Install_Tarrifs.append(Install_Tarrif)
        wb.create_sheet(title="Output")
        # ws_output.append([str(column_name)for column_name in ColName])
        # get all field in the row
        ws_output.append([row[i].value for i in range(len(row))])
        
#     for filterddata in Itinerary_Number:
#         # check if filtered data is intalled routes
#         if filterddata is inst_route:
#             # get the whole data for the filterddata
            
#             # row that has the value /filterddata
            
#             instat_tarrif=row[56].value
#             # print("instat_tarrif ",instat_tarrif)
            
#             if instat_tarrif is None:
#                 # print("instat_tarrif is None")
#                 meternumber=row[21].value
#                 # print("meternumber ",meternumber)
#                 accountNumber=row[36].value
#                 # print("accountNumber ",accountNumber)
                
                
#           wb.create_sheet(title="Output")
#           ws_output.append(["filterddata,","Meter Number", "Account Number"])
# #                 ws_output.append([filterddata,meternumber, accountNumber])
                
                



print(ws_output)

new_sheet = new_wb.active

# # Write the data to the new worksheet
for row_index, row_data in enumerate(ws_output, start=1):
    for col_index, value in enumerate(row_data, start=1):
        new_sheet.cell(row=row_index, column=col_index, value=value)

# # Rename the new worksheet
new_sheet.title = "Output Data"

# # Save the new workbook
new_wb.save('filtered_Data.xlsx')



